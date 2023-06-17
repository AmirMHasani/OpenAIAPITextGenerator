


from tkinter import *
from tkinter import filedialog
from PIL import Image, ImageTk, ImageDraw,ImageFont
import os,openpyxl,openai,datetime,json,copy,sys,pywintypes
from openpyxl.styles import Alignment,Font, Color, colors, Border, Side, PatternFill
from threading import Thread
from res.settings import settings_




WINDOW = Tk()

def progress_image(percentage=1):
    base_width = 300
    base_height = 30
    if float(percentage) > 100: percentage = 100
    percentage = int(percentage)
    percentage = percentage*3
    if percentage<= 35: percentage =35

    base_img = Image.new('RGBA',(base_width,base_height))
    canvas = ImageDraw.Draw(base_img)
    canvas.rounded_rectangle((5,5,base_width-5,base_height-5), radius=10, fill='#DDF2FF', outline='#97B2AF', width=0)
    if percentage> 35:
        canvas.rounded_rectangle((5,5,percentage-5,base_height-5), radius=10, fill='#00A89D', outline='#97B2AF', width=0)

    # base_img.show()
    return base_img

class gif():
    def __init__(self,LBL,image='',speed = 100,size=''):
        self.LBL = LBL
        self.speed = speed
        self.run_stop = True
        
        with Image.open(image) as im:
            sequence = []
            try:
                while True:
                    sequence.append(im.copy())
                    im.seek(im.tell() + 1)
            except EOFError:
                pass
        
        if type(size) is tuple :
            self.frames = [ImageTk.PhotoImage(img.resize((size[0],size[1]),Image.LANCZOS)) for img in sequence]
        else:
            self.frames = [ImageTk.PhotoImage(img) for img in sequence]
        self.LBL['image'] = self.frames[0]

    def start(self,frame=0):
        try: self.LBL.config(image=self.frames[frame])
        except: pass
        if self.run_stop == True:
            WINDOW.after(self.speed, self.start, (frame + 1) % len(self.frames))

    def stop(self):
        self.run_stop = False
        

def image_title(text='',height=30, width= 200, fontsize=18,style=1,color = 'black'):
    new_image = Image.new('RGBA',(width,height))  
    font = {}
    font[1] = ImageFont.truetype("res/monofonto_rg.otf", fontsize)
    font[2] = ImageFont.truetype("res/Molot.otf", fontsize)
    font[3] = ImageFont.truetype("res/Basketball.otf", fontsize)
    draw_image = ImageDraw.Draw(new_image)
    draw_image.text((int(width/2), height-5), text , fill=color, anchor="ms", font=font[style])
    width,height = new_image.size


    return new_image



class Evaluation():
    colorTk = {
        "w" : "white",
        "w1" : "#DFE5EE",
        "w3" : "#65666A",
        "g" : "#00733b",
        "g2" : "#656669",
        "tb" : "#f9f9f9",
        "b1" : "#326297",
    }
    def __init__(self,WINDOW,imageTK):
        self.WINDOW = WINDOW
        self.WINDOW['bg'] = self.colorTk['w']
        self.imageTK = imageTK

        self.WINDOW_HEIGHT = self.WINDOW.winfo_height()
        self.WINDOW_WIDTH = self.WINDOW.winfo_width()

        self.WINDOW.bind("<Configure>", self.on_window_resize_EL)

        #upload frame
        self.upload_frame = Frame(self.WINDOW,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.upload_frame.pack(anchor='w',fill=BOTH,expand=True)
        self.upload_frame_proparty()

        #working frame
        self.working_frame = Frame(self.WINDOW,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        # self.working_frame.pack(anchor='w')
        # self.working__frame_proparty()



        '''Ask for API key if previously not provided'''
        self.api_key_window_display = False
        self.api_key = ""
        if os.path.isfile('.apikey') :
            with open('.apikey','r') as RF:
                api_key = RF.read()
            RF.close()
            self.api_key = api_key.strip()

        if self.api_key == "":
            self.api_key_window_property()


        # Temporary
        # self.input_database_path = './input.xlsx'
        # self.output_database_path = "/".join(self.input_database_path.split('/')[:-1])+'/output_gpt_report.xlsx'
        # self.log_database_path = "/".join(self.input_database_path.split('/')[:-1])+'/gpt_report.log'

        # self.open_xlsx_file(self.input_database_path)

        self.data_curation_chat_list=[]
        self.data_curation_chat_list.append({"role": "system", "content": "Keep all answers in the same tone and structure. Always use correct grammar.Use medical language, terminology, and abbreviations. Avoid redundant and repetitive phrases and any information about series numbers. Ensure that the report is easy to read, organized, and includes all pertinent information. Avoid starting the output with a topic quotation."})
        self.data_standard_chat_list=[]
        self.data_standard_chat_list.append({"role": "system", "content": "Keep all answers in the same tone and structure. Always use correct grammar.Use medical language, terminology, and abbreviations. Avoid redundant and repetitive phrases and any information about series numbers. Ensure that the report is easy to read, organized, and includes all pertinent information. Avoid starting the output with a topic quotation."})
        self.data_keyfindings_chat_list=[]
        self.data_keyfindings_chat_list.append({"role": "system", "content": "Output must be in bullet point format. Keep all answers in the same tone and structure. Always use correct grammar. Avoid starting the output with a topic quotation."})
        
       



    def working__frame_proparty(self):
        self.run_all_reports = ''
        self.convertion_button_animation = False

        self.database_frame = Frame(self.working_frame,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0,padx=15)
        self.database_frame.pack()
        self.database_frame_property()

        self.title_frame = Frame(self.working_frame,border=0,borderwidth=0,highlightthickness=0)
        self.title_frame.pack()

        self.navigation_frame = Frame(self.working_frame,bg=self.colorTk['b1'],border=0,borderwidth=0,highlightthickness=0)
        self.navigation_frame.pack(anchor='e',fill=X,expand=True)
        self.navigation_frame_property()
        # 
        self.index_select_arrow_ILBL = Label(self.working_frame,bg=self.colorTk['w'],activebackground=self.colorTk['w'],width=32+77,border=0,borderwidth=0,highlightthickness=0)
        self.index_select_arrow_ILBL['image'] = self.imageTK['index_select']
        self.index_select_arrow_ILBL.image = self.imageTK['index_select']
        self.index_select_arrow_ILBL.pack(anchor='w')

        # top margin
        self.middle_frame_top_margin_ILBL = Label(self.working_frame,bg=self.colorTk['w'],activebackground=self.colorTk['w'],height=7,border=0,borderwidth=0,highlightthickness=0)
        self.middle_frame_top_margin_ILBL['image'] = self.imageTK['dot_margin']
        self.middle_frame_top_margin_ILBL.image = self.imageTK['dot_margin']
        self.middle_frame_top_margin_ILBL.pack(anchor='w')

        self.middle_frame = Frame(self.working_frame,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0,padx=15,pady=0)
        self.middle_frame.pack()
        self.middle_frame_property()
        
        # bottom margin
        self.middle_frame_bottom_margin_ILBL = Label(self.working_frame,bg=self.colorTk['w'],activebackground=self.colorTk['w'],height=15,border=0,borderwidth=0,highlightthickness=0)
        self.middle_frame_bottom_margin_ILBL['image'] = self.imageTK['dot_margin']
        self.middle_frame_bottom_margin_ILBL.image = self.imageTK['dot_margin']
        self.middle_frame_bottom_margin_ILBL.pack(anchor='w')

        
        self.log_frame  = Frame(self.working_frame,bg='#4B4453',border=0,borderwidth=0,highlightthickness=0,padx=15)
        self.log_frame.pack(anchor='w',expand=True,fill=BOTH)
        self.log_canvas  = Canvas(self.log_frame,bg='#4B4453',border=0,borderwidth=0,highlightthickness=0)
        self.log_canvas.pack(anchor='w',expand=True,fill=BOTH)
        self.log_text, self.log_canvas_base_text = self.make_scrollable_text(self.log_canvas,scrollbar=False,bg='#4B4453',fg='white')

        self.footer_frame  = Frame(self.working_frame,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.footer_frame.pack()
        self.footer_frame_property()

        
    def database_frame_property(self):
       

        self.file_location_display_frame = Canvas(self.database_frame,bg=self.colorTk['w'],height=50,border=0,borderwidth=0,highlightthickness=0)
        self.file_location_display_frame.pack(anchor='w',expand=True,fill=X)
        # Input file
        self.input_file_text_main_frame = Frame(self.file_location_display_frame,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.input_file_text_main_frame.pack(anchor='w',side=LEFT)

        self.input_file_text_ILBL = Label(self.input_file_text_main_frame,bg=self.colorTk['w'],activebackground=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.input_file_text_ILBL['image'] = self.imageTK['input_file']
        self.input_file_text_ILBL.image = self.imageTK['input_file']
        self.input_file_text_ILBL.pack(anchor='w')

        self.input_file_text_frame = Frame(self.input_file_text_main_frame,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.input_file_text_frame.pack(anchor='w')

        self.input_file_text_I_BTN = Label(self.input_file_text_frame,bg=self.colorTk['w'],activebackground=self.colorTk['w'],height=30,width=30,border=0,borderwidth=0,highlightthickness=0)
        self.input_file_text_I_BTN['image'] = self.imageTK['xls_24']
        self.input_file_text_I_BTN.image = self.imageTK['xls_24']
        self.input_file_text_I_BTN.pack(anchor='w',side=LEFT)

        self.input_file_text_display_Entry = Entry(self.input_file_text_frame,bg=self.colorTk['w'],fg=self.colorTk['g2'],font=('Arial','12','normal'),width=35,border=0,borderwidth=1,highlightthickness=0)
        self.input_file_text_display_Entry.pack(ipady=2,anchor='w',side=LEFT)
        self.input_file_text_display_Entry.insert(0,self.input_database_path)
        




        # output file
        self.output_file_text_main_frame = Frame(self.file_location_display_frame,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.output_file_text_main_frame.pack(side=LEFT)

        self.output_file_text_ILBL = Label(self.output_file_text_main_frame,bg=self.colorTk['w'],activebackground=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.output_file_text_ILBL['image'] = self.imageTK['output_file']
        self.output_file_text_ILBL.image = self.imageTK['output_file']
        self.output_file_text_ILBL.pack(anchor='w')

        self.output_file_text_frame = Frame(self.output_file_text_main_frame,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.output_file_text_frame.pack(anchor='w')

        self.output_file_text_I_BTN = Label(self.output_file_text_frame,bg=self.colorTk['w'],activebackground=self.colorTk['w'],height=30,width=30,border=0,borderwidth=0,highlightthickness=0)
        self.output_file_text_I_BTN['image'] = self.imageTK['xls_24']
        self.output_file_text_I_BTN.image = self.imageTK['xls_24']
        self.output_file_text_I_BTN.pack(anchor='w',side=LEFT)

        self.output_file_text_display_Entry = Entry(self.output_file_text_frame,bg=self.colorTk['w'],fg=self.colorTk['g2'],font=('Arial','12','normal'),width=35,border=0,borderwidth=1,highlightthickness=0)
        self.output_file_text_display_Entry.pack(ipady=2,anchor='w',side=LEFT)
        self.output_file_text_display_Entry.insert(0,self.output_database_path)


        # log file
        self.log_file_text_main_frame = Frame(self.file_location_display_frame,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.log_file_text_main_frame.pack(anchor='e',side=RIGHT)
         
        self.log_file_text_ILBL = Label(self.log_file_text_main_frame,bg=self.colorTk['w'],activebackground=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.log_file_text_ILBL['image'] = self.imageTK['log_file']
        self.log_file_text_ILBL.image = self.imageTK['log_file']
        self.log_file_text_ILBL.pack(anchor='w')

        self.log_file_text_frame = Frame(self.log_file_text_main_frame,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.log_file_text_frame.pack(anchor='w')

        self.log_file_text_I_BTN = Label(self.log_file_text_frame,bg=self.colorTk['w'],activebackground=self.colorTk['w'],height=30,width=30,border=0,borderwidth=0,highlightthickness=0)
        self.log_file_text_I_BTN['image'] = self.imageTK['log_24']
        self.log_file_text_I_BTN.image = self.imageTK['log_24']
        self.log_file_text_I_BTN.pack(anchor='w',side=LEFT)

        self.log_file_text_display_Entry = Entry(self.log_file_text_frame,bg=self.colorTk['w'],fg=self.colorTk['g2'],font=('Arial','12','normal'),width=35,border=0,borderwidth=1,highlightthickness=0)
        self.log_file_text_display_Entry.pack(ipady=2,anchor='w',side=LEFT)
        self.log_file_text_display_Entry.insert(0,self.log_database_path)
        
    


    def footer_frame_property(self):
        self.footer_frame_canvas = Canvas(self.footer_frame,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.footer_frame_canvas.pack(anchor='w',side=LEFT)

        self.footer_button_canvas =  Canvas(self.footer_frame_canvas,bg=self.colorTk['w3'],border=0,borderwidth=0,highlightthickness=0)
        self.footer_button_canvas.pack(anchor='w',side=LEFT)

        # settings button
        if not os.path.exists('res/data.log'):
            settings_(self.WINDOW)

        # self.chat_gpt_window = True

        self.settings_button = Button(self.footer_button_canvas,bg=self.colorTk['w3'],activebackground=self.colorTk['w3'],height=28,width=28,border=0,borderwidth=0,highlightthickness=0)
        self.settings_button['image'] = self.imageTK['settings']
        self.settings_button.image = self.imageTK['settings']
        self.settings_button['command'] = lambda window = self.WINDOW: settings_(window)
        self.settings_button.pack(anchor='w',side=LEFT) # 

        self.api_key_button = Button(self.footer_button_canvas,bg=self.colorTk['w3'],activebackground=self.colorTk['w3'],height=28,width=28,border=0,borderwidth=0,highlightthickness=0)
        self.api_key_button['image'] = self.imageTK['api_key']
        self.api_key_button.image = self.imageTK['api_key']
        self.api_key_button['command'] = self.api_key_window_property_BTN_Click
        self.api_key_button.pack(anchor='w',side=LEFT) 

        # log button
        self.log_text_show_status = False
        self.log_canvas_height = 0
        self.log_button = Button(self.footer_button_canvas,bg=self.colorTk['w3'],activebackground=self.colorTk['w3'],height=28,width=28,border=0,borderwidth=0,highlightthickness=0)
        self.log_button['image'] = self.imageTK['terminal']
        self.log_button.image = self.imageTK['terminal']
        self.log_button['command'] = self.show_hide_log_text
        self.log_button.pack(anchor='e',side=RIGHT)

    def show_hide_log_text(self):
        if self.log_text_show_status == False:
            self.log_text_show_status = True
            self.log_canvas_height = 90
            self.on_window_resize()
        else: 
            self.log_text_show_status = False
            self.log_canvas_height = 0
            self.on_window_resize()

    def log_write(self,data):

        #insert data at the last of the text field
        last_index = self.log_text.index('end')
        # print('------->',last_index)
        last_index = float(last_index) + 1
        # print('---->',last_index)
        self.log_text.insert(str(last_index),data)
        last_index = self.log_text.index('end')
        last_index = float(last_index) + 1
        self.log_text.see(str(last_index))

        with open(self.log_database_path,'w') as RF:
            RF.write(str(self.log_text.get('1.0',END)))
        RF.close()


    def api_key_window_property_BTN_Click(self):
        # print(self.api_key_window_display)
        if  self.api_key_window_display == True:
            self.api_key_window.focus_set()
        else: 
            self.api_key_window_property()

    
    def api_key_window_property(self):
        self.api_key_window_display = True
        self.api_key_window = Toplevel(self.WINDOW)
        self.api_key_window.attributes('-topmost',1)
        self.api_key_window['bg'] = self.colorTk['w']
        self.api_key_window.title('API Key')
        center_window(self.api_key_window,width=800,height=120)
 
        
        self.api_key_window.transient(self.WINDOW)
        self.api_key_window.protocol("WM_DELETE_WINDOW", self.save_api_key)
        self.api_key_window.iconphoto(True, self.imageTK['api_key'])
        self.api_key_window.resizable(0,0)
        # self.api_key_window.geometry('800x120')

        self.how_to_get_api_key_info = False

        self.api_key_Frame = Frame(self.api_key_window,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.api_key_Frame.pack()

        self.how_to_get_api_key_BTN = Button(self.api_key_Frame,text='► How to get api key?',font=('Arial','12','normal'),bg=self.colorTk['w'],activebackground=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.how_to_get_api_key_BTN['command'] = self.show_hide_how_to_get_api_key_info
        self.how_to_get_api_key_BTN.pack(anchor='w')

        

        self.api_instruction_frame = Canvas(self.api_key_Frame,width=577-5,bg=self.colorTk['w'],height=1,border=0,borderwidth=0,highlightthickness=0)
        self.api_instruction_frame.pack()

        self.api_instruction_internal_frame = Canvas(self.api_instruction_frame,width=577-5,bg=self.colorTk['w'],height=1,border=0,borderwidth=0,highlightthickness=0)
        # self.api_instruction_internal_frame.pack()

        self.api_instruction_url_frame = Frame(self.api_instruction_internal_frame,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.api_instruction_url_frame.pack()

        Label(self.api_instruction_url_frame,text='Go to ',bg=self.colorTk['w'],font=('Arial','12','normal')).pack(side=LEFT)
        
        self.api_url = 'https://platform.openai.com/account/api-keys'
        self.apiKey_url = Entry(self.api_instruction_url_frame,bg=self.colorTk['w'],font=('Arial','12','normal'),fg='#4E5BAD',width=35,border=0,borderwidth=0,highlightthickness=0)
        self.apiKey_url.pack(side=LEFT)
        self.apiKey_url.insert(0,'https://platform.openai.com/account/api-keys')
        self.apiKey_url.bind('<KeyRelease>',self.not_change_api_url)
        
        Label(self.api_instruction_url_frame,text=' and follow the instructions below.',bg=self.colorTk['w'],font=('Arial','12','normal')).pack(side=LEFT)

        api_instruction_GIFLBL = Label(self.api_instruction_internal_frame,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        gif(api_instruction_GIFLBL,image='res/api_guid.gif',speed = 100,size=(842-int(84.2*3),696-int(69.6*3))).start()
        api_instruction_GIFLBL.pack()

        self.api_key_Entry = Entry(self.api_key_Frame,bg=self.colorTk['w1'],fg=self.colorTk['g2'],font=('Arial','12','normal'),width=65,border=0,borderwidth=0,highlightthickness=0)
        self.api_key_Entry.pack(ipady=5,ipadx=2)
        self.api_key_Entry.delete(0,END)
        self.api_key_Entry.insert(0,self.api_key)
        self.api_key_Entry.bind('<KeyRelease>',self.save_api_key_EL)

        self.save_api_BTN = Button(self.api_key_Frame,bg=self.colorTk['w'],activeforeground=self.colorTk['w'],height=55,activebackground=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.save_api_BTN['image'] = self.imageTK['ok']
        self.save_api_BTN.image = self.imageTK['ok']
        self.save_api_BTN['command'] = self.save_api_key
        self.save_api_BTN.pack()

    def show_hide_how_to_get_api_key_info(self):
        if self.how_to_get_api_key_info == False:
            self.how_to_get_api_key_BTN['text'] = self.how_to_get_api_key_BTN['text'].replace('►','▼')
            self.how_to_get_api_key_info = True
            self.api_instruction_internal_frame.pack()
            self.api_instruction_frame['height'] = 400
            # self.api_key_window.geometry('800x660')
            center_window(self.api_key_window,width=800,height=660)
        else:
            self.how_to_get_api_key_info = False
            self.how_to_get_api_key_BTN['text'] = self.how_to_get_api_key_BTN['text'].replace('▼','►')
            self.api_instruction_frame['height'] = 1
            self.api_instruction_internal_frame.pack_forget()
            # self.api_key_window.geometry('800x120')
            center_window(self.api_key_window,width=800,height=120)

    def not_change_api_url(self,e):
         if self.api_url != self.apiKey_url.get().strip():
            self.apiKey_url.delete(0,END)
            self.apiKey_url.insert(0,self.api_url)

        
    def save_api_key_EL(self,e):
        self.api_key = self.api_key_Entry.get().strip()
        if self.api_key != '':
            with open ('.apikey','w') as RF:
                RF.write(self.api_key)
            RF.close()
            # print(self.api_key)

    def save_api_key(self):
        self.api_key = self.api_key_Entry.get().strip()
        if self.api_key != '':
            with open ('.apikey','w') as RF:
                RF.write(self.api_key)
            RF.close()
            # print(self.api_key)
            self.api_key_window_display = False
            self.api_key_window.destroy()


    def middle_frame_property(self):
        # LEFT FRAME
        # base text 


        self.middle_left_canvas = Canvas(self.middle_frame,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.middle_left_canvas.pack(anchor='w',side=LEFT)

        self.summary_report_LBL = Label(self.middle_left_canvas,bg=self.colorTk['w'],activebackground=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.summary_report_LBL['image'] = imageTK['original_report']
        self.summary_report_LBL.image = imageTK['original_report']
        self.summary_report_LBL.pack(anchor='w')

        self.base_text_canvas_1 = Canvas(self.middle_left_canvas,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.base_text_canvas_1.pack(anchor='w',side=LEFT)

        self.base_text_canvas = Canvas(self.base_text_canvas_1,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.base_text_canvas.pack(anchor='w')

        self.original_report_text, self.holding_canvas_base_text = self.make_scrollable_text(self.base_text_canvas)
        

        # COMMANDS
        self.commands_canvas = Canvas(self.base_text_canvas_1,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.commands_canvas.pack(anchor='w')

        self.list_of_commands_LBL = Label(self.commands_canvas,bg=self.colorTk['w'],activebackground=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.list_of_commands_LBL['image'] = imageTK['list_of_commands']
        self.list_of_commands_LBL.image = imageTK['list_of_commands']
        self.list_of_commands_LBL.pack(anchor='w')


        self.batch_run_button_frame = Frame(self.commands_canvas,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.batch_run_button_frame.pack(anchor='w')




        self.batch_run_CB = Checkbutton(self.batch_run_button_frame,bg=self.colorTk['w'],activebackground=self.colorTk['w'],variable = batch_run_boolVar,  onvalue = 1, offvalue = 0, height=0, width = 0,padx=0,pady=0 , command=self.batch_run_command_enable_disable)
        self.batch_run_CB.pack(anchor='w',side=LEFT)

        self.batch_run_LBL = Label(self.batch_run_button_frame,bg=self.colorTk['w'],activebackground=self.colorTk['w'],height=30,border=0,borderwidth=0,highlightthickness=0)
        self.batch_run_LBL['image'] = imageTK['run_report']
        self.batch_run_LBL.image = imageTK['run_report']
        self.batch_run_LBL.pack(anchor='w')


        self.data_command_LBL = Label(self.commands_canvas,bg=self.colorTk['w'],activebackground=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.data_command_LBL['image'] = imageTK['command_1']
        self.data_command_LBL.image = imageTK['command_1']
        self.data_command_LBL.pack(anchor='w')

        self.data_curation_command_Entry = Text(self.commands_canvas,height=3,bg=self.colorTk['tb'],fg=self.colorTk['g2'],font=('Arial','11','normal'),border=0,borderwidth=0,highlightthickness=0)
        self.data_curation_command_Entry.pack(ipady=0,anchor='w',expand=True,fill=X)



        self.key_finding_command_LBL = Label(self.commands_canvas,bg=self.colorTk['w'],activebackground=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.key_finding_command_LBL['image'] = imageTK['command_2']
        self.key_finding_command_LBL.image = imageTK['command_2']
        self.key_finding_command_LBL.pack(anchor='w')

        self.key_finding_command_Entry = Text(self.commands_canvas,height=3,bg=self.colorTk['tb'],fg=self.colorTk['g2'],font=('Arial','11','normal'),border=0,borderwidth=0,highlightthickness=0)
        # self.key_finding_command_Entry = 
        self.key_finding_command_Entry.pack(ipady=0,anchor='w',expand=True,fill=X)


        self.report_command_LBL = Label(self.commands_canvas,bg=self.colorTk['w'],activebackground=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.report_command_LBL['image'] = imageTK['command_3']
        self.report_command_LBL.image = imageTK['command_3']
        self.report_command_LBL.pack(anchor='w')

        self.report_standardize_command_Entry =  Text(self.commands_canvas,height=3,bg=self.colorTk['tb'],fg=self.colorTk['g2'],font=('Arial','11','normal'),border=0,borderwidth=0,highlightthickness=0)
        self.report_standardize_command_Entry.pack(ipady=0,anchor='w',expand=True,fill=X)
    




        # MIDDLE FRAME
        self.middle_middle_canvas = Canvas(self.middle_frame,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.middle_middle_canvas.pack(anchor='w',side=LEFT)
        # convert button
        self.convert_original_report_BTN = Button(self.middle_middle_canvas,bg=self.colorTk['w'],activebackground=self.colorTk['w'],width=65,border=0,borderwidth=0,highlightthickness=0)
        self.convert_original_report_BTN['image'] = self.imageTK['generate']
        self.convert_original_report_BTN.image = self.imageTK['generate']
        self.convert_original_report_BTN['command'] = self.start_converting
        self.convert_original_report_BTN.pack(anchor='w',expand=True,fill=Y,side=LEFT)

        self.convert_original_report_animation_BTN = Button(self.middle_middle_canvas,bg=self.colorTk['w'],activebackground=self.colorTk['w'],width=65,border=0,borderwidth=0,highlightthickness=0)
        # self.convert_original_report_animation_BTN['image'] = self.imageTK['generate']
        # self.convert_original_report_animation_BTN.image = self.imageTK['generate']
        # self.convert_original_report_animation_BTN['command'] = self.start_converting
        self.convert_original_report_animation_BTN['relief'] = SUNKEN
        self.convert_original_report_animation_BTN['command'] = ''
        self.convert_animation = gif(self.convert_original_report_animation_BTN,image='res/generate.gif',speed=65,size=(60,60))
        self.convert_animation.start()
        # self.convert_original_report_animation_BTN.pack(anchor='w',expand=True,fill=Y,side=LEFT)









        # RIGHT FRAME
        self.middle_right_canvas = Canvas(self.middle_frame,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.middle_right_canvas.pack(anchor='w')

        # Standarize report
        self.standarize_report_canvas =Canvas(self.middle_right_canvas,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.standarize_report_canvas.pack(anchor='w')
        # LBL
        self.Standarize_report_LBL = Label(self.standarize_report_canvas,bg=self.colorTk['w'],activebackground=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.Standarize_report_LBL['image'] = imageTK['std_report']
        self.Standarize_report_LBL.image = imageTK['std_report']
        self.Standarize_report_LBL.pack(anchor='w')
        # Loading
        self.Standarize_report_LOODING_canvas =Canvas(self.standarize_report_canvas,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        # self.Standarize_report_LOODING_canvas.pack()
        self.standarize_report_loading_ILBL = Label(self.Standarize_report_LOODING_canvas,bg=self.colorTk['tb'],border=0,borderwidth=0,highlightthickness=1)
        self.standarize_report_loading_ILBL.pack(side=LEFT)
        gif(self.standarize_report_loading_ILBL,image='res/hourglass.gif',speed=100,size=(128,128)).start()
        # Margin
        # Label(self.Standarize_report_LOODING_canvas,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0,padx=7).pack(anchor='w',side=LEFT)

       

        self.standarize_report_text_canvas = Canvas(self.standarize_report_canvas,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.standarize_report_text_canvas.pack()
        self.standarize_report_text , self.holding_frame_standarize_report= self.make_scrollable_text(self.standarize_report_text_canvas)


              
        # Key findings
        self.key_findingds_canvas =Canvas(self.middle_right_canvas,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.key_findingds_canvas.pack(anchor='w')
        # Lable
        self.key_findings_summary_report_LBL = Label(self.key_findingds_canvas,bg=self.colorTk['w'],activebackground=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.key_findings_summary_report_LBL['image'] = imageTK['key_findings']
        self.key_findings_summary_report_LBL.image = imageTK['key_findings']
        self.key_findings_summary_report_LBL.pack(anchor='w')
        # Loading
        self.key_findingds_LOODING_canvas =Canvas(self.key_findingds_canvas,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        # self.key_findingds_LOODING_canvas.pack()
        self.key_findings_loading_ILBL = Label(self.key_findingds_LOODING_canvas,bg=self.colorTk['tb'],border=0,borderwidth=0,highlightthickness=1)
        self.key_findings_loading_ILBL.pack()
        gif(self.key_findings_loading_ILBL,image='res/hourglass.gif',speed=100,size=(128,128)).start()
       

        self.key_findingds_text_canvas = Canvas(self.key_findingds_canvas,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.key_findingds_text_canvas.pack(anchor='w')
        self.key_findings_text , self.holding_frame_key_findings = self.make_scrollable_text(self.key_findingds_text_canvas)
       



       
        
    def batch_run_command_enable_disable(self):
        if batch_run_boolVar.get() == 1:
            self.batch_run_LBL['image'] = imageTK['run_report2']
            self.batch_run_LBL.image = imageTK['run_report2']
        else:
            self.batch_run_LBL['image'] = imageTK['run_report']
            self.batch_run_LBL.image = imageTK['run_report']
        
        



        
        
        


    def make_scrollable_text(self,canvas,scrollbar=True,bg='',fg=''):
        if bg == '':
            bg=self.colorTk['tb']
        if fg == '':
            fg=self.colorTk['g2']
        
        # canvas_new = Canvas(canvas,border=0,borderwidth=0,highlightthickness=0)
        # canvas_new.pack()
        # base passage
        text_widget = Text(canvas,bg=bg,fg=fg,font=('Arial','11','normal'),border=0,borderwidth=0,highlightthickness=0)
        # Scroll bar
        if scrollbar == True:
            scrollbar = Scrollbar(canvas)
            scrollbar.pack(side=RIGHT, fill=Y)
            text_widget.config(yscrollcommand=scrollbar.set)
            scrollbar.config(command=text_widget.yview)

        text_widget.pack(anchor='w',expand=True,fill=BOTH)

        return text_widget,canvas

    def start_converting(self):
        print('START FROM HERE')
        #Check batch run 
        # self.run_all_reports = ''
        # Batch run True
        if batch_run_boolVar.get() == True:
            # Check any job is progressed
            if self.total_processed_report != 0 :
                self.check_batch_run_mode()
            else:
                # self.batch_run_report()
                threadFun = Thread(target=self.batch_run_report)
                threadFun.start()    
        else:
            if self.root_data_process_dict[self.selected_report_index] == True:
                self.total_processed_report -= 1
                self.index_image()
            threadFun = Thread(target=self.convert_by_gpt,args=(self.selected_report_index,))
            threadFun.start()
        


        
    def batch_run_report(self):
        # Check run for all reports and overwrite
        if batch_run_boolVar.get() == True:
            self.total_processed_report = 0
            self.index_image()
    
        print('--*',batch_run_boolVar.get())
        self.batch_run_count = 0
        threadFun = {}

        self.total_batch_report = 0
        self.finished_batch_report = 0
        for each_process in self.root_data_process_dict:
            if batch_run_boolVar.get() == False :
                if self.root_data_process_dict[each_process] == False:
                    self.total_batch_report += 1     
            else:
                self.total_batch_report += 1

        for each_process in self.root_data_process_dict:
            if batch_run_boolVar.get() == False :
                if self.root_data_process_dict[each_process] == False:
                    self.convert_by_gpt(each_process)

            else:
                self.root_data_process_dict[each_process] = False
                self.convert_by_gpt(each_process)  

        


    def check_batch_run_mode(self):
        self.batch_run_message_window = Toplevel(self.WINDOW)
        center_window(self.batch_run_message_window,width=400,height=170)
        self.batch_run_message_window['bg'] = 'white'
        self.batch_run_message_window.resizable(0,0)
        self.batch_run_message_window.protocol("WM_DELETE_WINDOW", lambda : True)
        self.batch_run_message_window.attributes('-toolwindow',1)
        # self.WINDOW.attributes('-topmost',1)
        self.batch_run_message_window.attributes('-topmost',1)
        self.batch_run_message_window.title('Report Generation Selection')
        self.batch_run_message_window.geometry('400x170')
        app_icon = Image.open("res/logo.png")
        app_icon = ImageTk.PhotoImage(app_icon.resize((16,16),Image.LANCZOS))
        self.batch_run_message_window.iconphoto(False, app_icon)

        Label(self.batch_run_message_window,bg= 'white',font=('Arial','12','bold'),padx=30).pack()

        Label(self.batch_run_message_window,bg= 'white',text='Do you want to overwrite all the report?',font=('Arial','12','bold')).pack()
        Label(self.batch_run_message_window,bg= 'white',text="Select 'Yes' to run all the reports or 'No' to run the remaining unprocessed reports.",fg='#6c757d',justify=LEFT, wraplength=310,font=('Arial','10','normal')).pack()

        Label(self.batch_run_message_window,bg= 'white',font=('Arial','12','bold'),padx=30).pack()
        button_frame = Frame(self.batch_run_message_window,bg= 'white',border=0,borderwidth=0,highlightthickness=0)
        button_frame.pack()

        yes_button = Button(button_frame,text='Yes',bg='#52b788',fg= 'white',activebackground='#74c69d',activeforeground='#dee2e6',width=7,font=('Arial','12','bold'),border=0,borderwidth=0,highlightthickness=0)
        yes_button['command'] = lambda : self.on_click_bach_report_mood(True)
        yes_button.pack(side=LEFT)
        Label(button_frame,bg= 'white',font=('Arial','12','bold'),padx=50).pack(side=LEFT)
        no_button = Button(button_frame,text='No',bg='#e76f51',fg= 'white',activebackground='#EA8168',activeforeground='#dee2e6',width=7,font=('Arial','12','bold'),border=0,borderwidth=0,highlightthickness=0)
        no_button['command'] = lambda : self.on_click_bach_report_mood(False)
        no_button.pack(side=LEFT)
    def on_click_bach_report_mood(self,mode):
        # self.run_all_reports = mode
        self.batch_run_message_window.destroy()
        # self.batch_run_report()

        threadFun = Thread(target=self.batch_run_report)
        threadFun.start()   
      

    def navigation_frame_property(self):


        # Number
        self.original_report_selected_index_ILBL = Label(self.navigation_frame,bg=self.colorTk['b1'],activebackground=self.colorTk['b1'],border=0,borderwidth=0,highlightthickness=0)
        self.original_report_selected_index_ILBL.pack(anchor='w',side=LEFT)

        self.progress_bar_frame = Frame(self.navigation_frame,bg=self.colorTk['b1'], border=0,borderwidth=0,highlightthickness=0)
        self.progress_bar_frame.pack(anchor='center',side=LEFT)

        self.progress_bar_ILB = Label(self.progress_bar_frame,bg=self.colorTk['b1'],height=50,activebackground=self.colorTk['b1'],border=0,borderwidth=0,highlightthickness=0)
        self.progress_bar_ILB.pack(anchor='w',side=LEFT)
        

        #index_image(self,current) 
        self.processed_report_number_display_ILBL = Label(self.progress_bar_frame,bg=self.colorTk['b1'],activebackground=self.colorTk['b1'],border=0,borderwidth=0,highlightthickness=0)
        self.processed_report_number_display_ILBL.pack(anchor='w',side=LEFT)



        self.naviagtion_property_frame =Frame(self.navigation_frame,bg=self.colorTk['b1'], border=0,borderwidth=0,highlightthickness=0)
        self.naviagtion_property_frame.pack(anchor='e',side=RIGHT)
        # Margin
        Label(self.naviagtion_property_frame,bg=self.colorTk['b1'],border=0,borderwidth=0,highlightthickness=0,padx=7).pack(anchor='w',side=LEFT)

        # Left arrow 
        self.left_shift_BTN = Button(self.naviagtion_property_frame,bg=self.colorTk['b1'],height=50,width=45,activebackground=self.colorTk['b1'],border=0,borderwidth=0,highlightthickness=0)
        self.left_shift_BTN['image'] = imageTK['left_gray']
        self.left_shift_BTN.image = imageTK['left_gray']
        self.left_shift_BTN['relief'] =  SUNKEN
        # self.left_shift_BTN['command'] = lambda : self.shift_report(type='-')
        self.left_shift_BTN.pack(anchor='w',side=LEFT)

        # Margin
        Label(self.naviagtion_property_frame,bg=self.colorTk['b1'],border=0,borderwidth=0,highlightthickness=0,padx=7).pack(anchor='w',side=LEFT)

        # Right arrow
        self.right_shift_BTN = Button(self.naviagtion_property_frame,bg=self.colorTk['b1'],height=50,width=45,activebackground=self.colorTk['b1'],border=0,borderwidth=0,highlightthickness=0)
        self.right_shift_BTN['image'] = imageTK['right']
        self.right_shift_BTN.image = imageTK['right']
        self.right_shift_BTN['command'] = lambda : self.shift_report(type='+')
        self.right_shift_BTN.pack(anchor='w',side=LEFT)

        # Margin
        Label(self.naviagtion_property_frame,bg=self.colorTk['b1'],border=0,borderwidth=0,highlightthickness=0,padx=7).pack(anchor='w',side=LEFT)

        
        



    '''This method responsible for swip between reports'''
    def shift_report(self,type=''):
        
        # FORWARD
        if type == '+':
            if self.selected_report_index + 1 <= self.total_row:
                self.selected_report_index += 1
                if self.selected_report_index < self.total_row:
                    #enable pin
                    self.right_shift_BTN['image'] = imageTK['right']
                    self.right_shift_BTN.image = imageTK['right']
                    self.right_shift_BTN['command'] = lambda : self.shift_report(type='+')
                    self.right_shift_BTN['relief'] =  RAISED

                    if self.selected_report_index > 1:
                        self.left_shift_BTN['image'] = imageTK['left']
                        self.left_shift_BTN.image = imageTK['left']
                        self.left_shift_BTN['command'] = lambda : self.shift_report(type='-')
                        self.left_shift_BTN['relief'] =  RAISED
                else:
                    #disable pin
                    self.right_shift_BTN['image'] = imageTK['right_gray']
                    self.right_shift_BTN.image = imageTK['right_gray']
                    self.right_shift_BTN['command'] = ''
                    self.right_shift_BTN['relief'] =  SUNKEN

                    self.left_shift_BTN['image'] = imageTK['left']
                    self.left_shift_BTN.image = imageTK['left']
                    self.left_shift_BTN['command'] = lambda : self.shift_report(type='-')
                    self.left_shift_BTN['relief'] =  RAISED


        # BACKWARD
        elif type == '-':
            if self.selected_report_index - 1 >= 1:
                self.selected_report_index -= 1
                if self.selected_report_index > 1:
                    #enable pin
                    self.left_shift_BTN['image'] = imageTK['left']
                    self.left_shift_BTN.image = imageTK['left']
                    self.left_shift_BTN['command'] = lambda : self.shift_report(type='-')
                    self.left_shift_BTN['relief'] =  RAISED

                    if self.selected_report_index < self.total_row:
                        self.right_shift_BTN['image'] = imageTK['right']
                        self.right_shift_BTN.image = imageTK['right']
                        self.right_shift_BTN['command'] = lambda : self.shift_report(type='+')
                        self.right_shift_BTN['relief'] =  RAISED        
                else:
                    #disable pin
                    self.left_shift_BTN['image'] = imageTK['left_gray']
                    self.left_shift_BTN.image = imageTK['left_gray']
                    self.left_shift_BTN['command'] = ''
                    self.left_shift_BTN['relief'] =  SUNKEN

                    self.right_shift_BTN['image'] = imageTK['right']
                    self.right_shift_BTN.image = imageTK['right']
                    self.right_shift_BTN['command'] = lambda : self.shift_report(type='+')
                    self.right_shift_BTN['relief'] =  RAISED

        # print(self.selected_report_index,self.total_row)           
        # DISPLAY SELECTED INDEX 
        self.current_index()
        # DISPLAY SELECTED INDEX :: ORIGINAL REPORT
        self.original_report_text.delete(1.0,END)
        self.original_report_text.insert(1.0,self.root_data_dict[self.selected_report_index]["original_report"])

        self.standarize_report_text.delete(1.0,END)
        self.standarize_report_text.insert(1.0,self.root_data_dict[self.selected_report_index]["standardize_report"])

        self.key_findings_text.delete(1.0,END)
        self.key_findings_text.insert(1.0,self.root_data_dict[self.selected_report_index]["key_findings"])






    '''This method call the open ai API '''
    def openai_api_call(self,gpt_instruction='',chat_gpt_configure=[] ,type__=''):
        gpt_instruction = gpt_instruction+'\n\n\n'
        # print(json.dumps(chat_gpt_configure,indent=4))
        # print('v---INPUT---v'*5,'\n',gpt_instruction)

        # Complete
        openai.api_key = self.api_key

        chat_model_list = ['gpt-4','gpt-3.5-turbo','gpt-3.5-turbo-0301']
        if chat_gpt_configure['model'] in chat_model_list:

            all_chat_message = []
            input_message = {"role": "user", "content": gpt_instruction}

            
            if type__=="KEY FINDINGS":
                if len(self.data_keyfindings_chat_list) == 1:
                    # print('\n------->',self.data_keyfindings_chat_list,'<---------\n\n')
                    self.data_keyfindings_chat_list.append(input_message)
                    all_chat_message= copy.deepcopy(self.data_keyfindings_chat_list)
                else:
                    all_chat_message = copy.deepcopy(self.data_keyfindings_chat_list)
                    all_chat_message.append(input_message)
            
            elif type__=="STANDARIZE REPORT":
                if len(self.data_standard_chat_list) == 1:
                    # print('\n------->',self.data_standard_chat_list,'<---------\n\n')
                    self.data_standard_chat_list.append(input_message)
                    all_chat_message = copy.deepcopy(self.data_standard_chat_list)
                else:
                    all_chat_message = copy.deepcopy(self.data_standard_chat_list)
                    all_chat_message.append(input_message)

            elif type__=="DATA CURATION":
                if len(self.data_curation_chat_list) == 1:
                    # print('\n------->',self.data_curation_chat_list,'<---------\n\n')
                    self.data_curation_chat_list.append(input_message)
                    all_chat_message= copy.deepcopy(self.data_curation_chat_list)
                else:
                    all_chat_message = copy.deepcopy(self.data_curation_chat_list)
                    all_chat_message.append(input_message)
            
           

                
            
            print('\n\n\n===>',datetime.datetime.now(),type__,len(all_chat_message))
            print('\nINPUT:',json.dumps(all_chat_message,indent=4))

            response = openai.ChatCompletion.create(\
                model=chat_gpt_configure['model'], \
                temperature = float(chat_gpt_configure['temperature']),\
                max_tokens = int(chat_gpt_configure['max_tokens']),\
                top_p = float(chat_gpt_configure['top_p']),\
                frequency_penalty = float(chat_gpt_configure['frequency_penalty']),\
                presence_penalty = float(chat_gpt_configure['presence_penalty']),\
                messages= all_chat_message\
                )
            
            # print('\n\n\n===>',response)
            # response = str(response[ "choices"][0]["message"]["content"]).replace('Key Findings:','').strip()
            response = str(response[ "choices"][0]["message"]["content"]).strip()
            # print('\n\n\n===>',json.dumps(all_chat_message,indent=4))


            response_message = {"role": "assistant", "content": str(response)}
            
            if type__=="KEY FINDINGS":
                if len(self.data_keyfindings_chat_list) == 2:
                    # self.data_keyfindings_chat_list.append(response_message)
                    self.data_keyfindings_chat_list[-1] = response_message
            
            elif type__=="STANDARIZE REPORT":
                if len(self.data_standard_chat_list) == 2:
                    # self.data_standard_chat_list.append(response_message)
                    self.data_standard_chat_list[-1] = response_message

            elif type__=="DATA CURATION":
                if len(self.data_curation_chat_list) == 2:
                    # self.data_curation_chat_list.append(response_message)
                    self.data_curation_chat_list[-1] = response_message

            all_chat_message.append(response_message)
            print('\nOUTPUT:',json.dumps(all_chat_message,indent=4))


        else:
            
            response = openai.Completion.create(\
                model=chat_gpt_configure['model'], \
                temperature=float(chat_gpt_configure['temperature']),\
                max_tokens=int(chat_gpt_configure['max_tokens']),\
                top_p=float(chat_gpt_configure['top_p']),\
                best_of=int(chat_gpt_configure['best_of']),\
                frequency_penalty=float(chat_gpt_configure['frequency_penalty']),\
                presence_penalty=float(chat_gpt_configure['presence_penalty']),\
                prompt= str(gpt_instruction)   )
            
            response = str(response[ "choices"][0]["text"]).replace('Key Findings:','').strip()

        return response




    '''This method is responsible for generating the data through openai API'''
    def convert_by_gpt(self,index_number):
        # print('--------------------------------\n',json.dumps(self.root_data_dict[index_number],indent=4))
        if self.convertion_button_animation == False:
            self.convertion_button_animation = True
            self.convert_original_report_BTN.pack_forget()
            self.convert_original_report_animation_BTN.pack(anchor='w',expand=True,fill=Y,side=LEFT)

            # print(self.left_shift_BTN['state'])
            self.left_shift_BTN['state'] = DISABLED
            self.right_shift_BTN['state'] = DISABLED
            # print(self.left_shift_BTN['state'])

        standardization_command = '''
Make a standardized with below criteria on below report
Use a standardized format appropriate for an abdominal radiology report in a clinical setting. 
Use medical language, terminology, and abbreviations.
Avoid redundant and repetitive phrases and any information about series numbers.
Ensure that the report is easy to read, organized, and includes all pertinent information.
        '''
        keyfindings_command = '''
Extract the key findings of the following information and list it in bullet point format:
        '''
        
        # print('index_number : ',index_number)
        # Read GPT Configure
        f = open('res/data.log')
        chat_gpt_configure = json.load(f)
        
        # Show standarize report loading
        self.standarize_report_text_canvas.pack_forget()
        self.Standarize_report_LOODING_canvas.pack()
        # Show key finding report loading
        self.key_findingds_text_canvas.pack_forget()
        self.key_findingds_LOODING_canvas.pack()


        # DATA CURATION
        # DATA CURATION :: START :: LOG  
        data = f'{str(datetime.datetime.now().strftime("%I:%M:%S %p")).ljust(16," ")}{str(index_number).ljust(10," ")}{str("DATA CURATION").ljust(40," ")}{str("SUBMITTED").ljust(15," ")}\n'
        self.log_write(data)
        # DATA CURATION :: GENERATE 
        try:
            if len(str(self.data_curation_command_Entry.get('1.0',END)).strip()) > 0:
                    gpt_instruction = str(self.data_curation_command_Entry.get('1.0',END)).strip() +'\n\n'+ self.root_data_dict[index_number]["original_report"]+'\n\n\n'
            else:
                gpt_instruction = '???????' +'\n\n'+ self.root_data_dict[index_number]["original_report"]+'\n\n\n'
            # print('[1]=>\n',gpt_instruction)
            response = self.openai_api_call(chat_gpt_configure=chat_gpt_configure,gpt_instruction=gpt_instruction,type__='DATA CURATION')
                
            # DATA CURATION :: END :: LOG 
            data = f'{str(datetime.datetime.now().strftime("%I:%M:%S %p")).ljust(16," ")}{str(index_number).ljust(10," ")}{str("DATA CURATION").ljust(40," ")}{str("RETURNED").ljust(15," ")}\n'
            self.log_write(data)

            # DATA CURATION :: SAVE TO MAIN DICT 
            # print('\n\nRESPONSE :: ',response)
            self.root_data_dict[index_number]["data_curation"] = response.strip()
            self.root_data_dict[index_number]["data_curation_command"] = str(self.data_curation_command_Entry.get('1.0',END))

        except Exception as e:
            # DATA CURATION :: OPENAI ERROR 
            data = f'\n{str(datetime.datetime.now().strftime("%I:%M:%S %p")).ljust(16," ")}\t\t{str(e)}\n'
            self.log_write(data)
        

        

        

        # STANDARIZE REPORT
        # STANDARIZE REPORT :: START :: LOG  
        data = f'{str(datetime.datetime.now().strftime("%I:%M:%S %p")).ljust(16," ")}{str(index_number).ljust(10," ")}{str("STANDARDIZE REPORT").ljust(40-8," ")}{str("SUBMITTED").ljust(15," ")}\n'
        self.log_write(data)

        # STANDARIZE REPORT :: GENERATE 
        try:
            if len(str(self.report_standardize_command_Entry.get('1.0',END)).strip()) > 0:
                    gpt_instruction = str(self.report_standardize_command_Entry.get('1.0',END)).strip() +'\n\n'+ self.root_data_dict[index_number]["original_report"]+'\n\n\n'
            else:
                gpt_instruction = standardization_command +'\n\n'+ self.root_data_dict[index_number]["original_report"]+'\n\n\n'
            # print('[1]=>\n',gpt_instruction)
            response = self.openai_api_call(chat_gpt_configure=chat_gpt_configure,gpt_instruction=gpt_instruction,type__='STANDARIZE REPORT')
                
            # STANDARIZE REPORT :: END :: LOG 
            data = f'{str(datetime.datetime.now().strftime("%I:%M:%S %p")).ljust(16," ")}{str(index_number).ljust(10," ")}{str("STANDARDIZE REPORT").ljust(40-8," ")}{str("RETURNED").ljust(15," ")}\n'
            self.log_write(data)

            # STANDARIZE REPORT :: SAVE TO MAIN DICT 
            # print('\n\nRESPONSE :: ',response)
            self.root_data_dict[index_number]["standardize_report"] = response.strip()
            self.root_data_dict[index_number]["standardize_report_command"] = str(self.report_standardize_command_Entry.get('1.0',END))

        except Exception as e:
            # STANDARIZE REPORT :: OPENAI ERROR 
            data = f'\n{str(datetime.datetime.now().strftime("%I:%M:%S %p")).ljust(16," ")}\t\t{str(e)}\n'
            self.log_write(data)
 
        
        if self.selected_report_index == index_number:            
            # STANDARIZE REPORT :: DISPLAY
            self.standarize_report_text.delete(1.0,END)
            self.standarize_report_text.insert(1.0,self.root_data_dict[index_number]["standardize_report"])
            pass

        # STANDARIZE REPORT :: HIDE LOADING 
        if batch_run_boolVar.get() == False:
            self.Standarize_report_LOODING_canvas.pack_forget()
            self.standarize_report_text_canvas.pack()
        else: 
            if self.finished_batch_report+1 == self.total_batch_report :
                self.Standarize_report_LOODING_canvas.pack_forget()
                self.standarize_report_text_canvas.pack()

        

        

        # KEY FINDINGS
        # KEY FINDINGS :: START :: LOG  
        data = f'{str(datetime.datetime.now().strftime("%I:%M:%S %p")).ljust(16," ")}{str(index_number).ljust(10," ")}{str("KEY FINDING").ljust(40+4," ")}{str("SUBMITTED").ljust(15," ")}\n'
        self.log_write(data)

        # KEY FINDINGS :: GENERATE 
        try:

            # prompt=f"{str(self.key_finding_command_Entry.get('1.0',END))} \n"+ self.original_report_text.get(0.0,END)    )
            if len(str(self.report_standardize_command_Entry.get('1.0',END)).strip()) > 0:
                    gpt_instruction = str(self.key_finding_command_Entry.get('1.0',END)).strip() +'\n\n'+ self.root_data_dict[index_number]["original_report"]+'\n\n\n'
            else:
                gpt_instruction = keyfindings_command +'\n\n'+ self.root_data_dict[index_number]["original_report"]+'\n\n\n'
            # print('=>\n',gpt_instruction)
            response = self.openai_api_call(chat_gpt_configure=chat_gpt_configure,gpt_instruction=gpt_instruction,type__='KEY FINDINGS')
            
            # KEY FINDINGS :: END :: LOG 
            data = f'{str(datetime.datetime.now().strftime("%I:%M:%S %p")).ljust(16," ")}{str(index_number).ljust(10," ")}{str("KEY FINDING").ljust(40+4," ")}{str("RETURNED").ljust(15," ")}\n'
            self.log_write(data)

            # KEY FINDINGS :: SAVE TO MAIN DICT
            # print('\n\nRESPONSE :: ',response)
            self.root_data_dict[index_number]["key_findings"] = response
            self.root_data_dict[index_number]["key_findings_command"] = str(self.key_finding_command_Entry.get('1.0',END))
        except Exception as e:

            # KEY FINDINGS :: OPENAI ERROR
            data = f'\n{str(datetime.datetime.now().strftime("%I:%M:%S %p")).ljust(16," ")}\t\t{str(e)}\n'
            self.log_write(data)

        
        if self.selected_report_index == index_number:
            # KEY FINDINGS :: DISPLAY
            self.key_findings_text.delete(0.0,END)
            self.key_findings_text.insert(0.0,self.root_data_dict[index_number]["key_findings"])

        # KEY FINDINGS:: HIDE LOADING 
        if batch_run_boolVar.get() == False:
            self.key_findingds_LOODING_canvas.pack_forget()
            self.key_findingds_text_canvas.pack()
        else: 
            if self.finished_batch_report+1 == self.total_batch_report :
                self.key_findingds_LOODING_canvas.pack_forget()
                self.key_findingds_text_canvas.pack()
        
        
        # UPDATE 
        self.root_data_process_dict[index_number] = True
        
        
        
        # UPDATE PROGRESSBAR
        self.total_processed_report += 1
        self.index_image()

        # UPDATE CONVERT BUTTON
        if batch_run_boolVar.get() == True:
            self.finished_batch_report += 1
            if self.finished_batch_report == self.total_batch_report :
                self.convert_original_report_animation_BTN.pack_forget()
                self.convert_original_report_BTN.pack(anchor='w',expand=True,fill=Y,side=LEFT)

                self.left_shift_BTN['state'] = NORMAL
                self.right_shift_BTN['state'] = NORMAL

                self.convertion_button_animation = False
        else:
            self.convert_original_report_animation_BTN.pack_forget()
            self.convert_original_report_BTN.pack(anchor='w',expand=True,fill=Y,side=LEFT)

            self.left_shift_BTN['state'] = NORMAL
            self.right_shift_BTN['state'] = NORMAL

            self.convertion_button_animation = False

        # LOG BREAK
        data='\n'
        self.log_write(data)

        # SCROLL LAST LOG 
        self.log_text.see(END)

        # SAVE DATA REAL TIME
        self.save_output_file_RT()


    ''' This method is responsidle to create output excel file'''      
    def save_output_file_RT(self):
        output_excel_file = openpyxl.Workbook()
        sheet = output_excel_file.active
        # Sheet Title
        sheet.title = "Report"

        # Border Style
        border = Border(left=Side(border_style='thin', color='adb5bd'),
                right=Side(border_style='thin', color='adb5bd'),
                top=Side(border_style='thin', color='adb5bd'),
                bottom=Side(border_style='thin', color='adb5bd'))
         
        # Column width
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 63
        sheet.column_dimensions['C'].width = 63
        sheet.column_dimensions['D'].width = 63
        sheet.column_dimensions['E'].width = 63
        sheet.column_dimensions['F'].width = 63

        # Column Heading
        sheet.cell(row = 1, column = 1).value = 'ID'
        sheet['A1'].font = Font(bold=True)
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet['A1'].border = border
        
        sheet.cell(row = 1, column = 2).value = 'Original Report'
        sheet['B1'].font = Font(bold=True)
        sheet['B1'].alignment = Alignment(horizontal='center')
        sheet['B1'].border = border

        sheet.cell(row = 1, column = 3).value = 'Data curation'
        sheet['C1'].font = Font(bold=True)
        sheet['C1'].alignment = Alignment(horizontal='center')
        sheet['C1'].border = border

        sheet.cell(row = 1, column = 4).value = 'Key Findings'
        sheet['D1'].font = Font(bold=True)
        sheet['D1'].alignment = Alignment(horizontal='center')
        sheet['D1'].border = border

        sheet.cell(row = 1, column = 5).value = 'Standardized report'
        sheet['E1'].font = Font(bold=True)
        sheet['E1'].alignment = Alignment(horizontal='center')
        sheet['E1'].border = border
        
        # Place Data 
        dala_of_column = {}
        row = 1
        for each_report in self.root_data_dict:
            row += 1
            dala_of_column['A'] = sheet.cell(row = row, column = 1)
            dala_of_column['A'] .value = str(self.root_data_dict[each_report]['report_id'])
            dala_of_column['A'].alignment = Alignment(wrap_text=True,vertical='top')
            dala_of_column['A'].fill = PatternFill("solid", start_color="cbf3f0")
            dala_of_column['A'].border = border

            dala_of_column['B'] = sheet.cell(row = row, column = 2)
            dala_of_column['B'].value = str(self.root_data_dict[each_report]['original_report'])
            dala_of_column['B'].alignment = Alignment(wrap_text=True,vertical='top')
            dala_of_column['B'].fill = PatternFill("solid", start_color="d8f3dc")
            dala_of_column['B'].border = border

            dala_of_column['C'] = sheet.cell(row = row, column = 3)
            dala_of_column['C'].value = str(self.root_data_dict[each_report]['data_curation'])
            dala_of_column['C'].alignment = Alignment(wrap_text=True,vertical='top')
            dala_of_column['C'].fill = PatternFill("solid", start_color="b7e4c7")
            dala_of_column['C'].border = border

            dala_of_column['D'] = sheet.cell(row = row, column = 4)
            dala_of_column['D'].value = str(self.root_data_dict[each_report]['key_findings'])
            dala_of_column['D'].alignment = Alignment(wrap_text=True,vertical='top')
            dala_of_column['D'].fill = PatternFill("solid", start_color="d8f3dc")
            dala_of_column['D'].border = border

            dala_of_column['E'] = sheet.cell(row = row, column = 5)
            dala_of_column['E'].value = str(self.root_data_dict[each_report]['standardize_report'])
            dala_of_column['E'].alignment = Alignment(wrap_text=True,vertical='top')
            dala_of_column['E'].fill = PatternFill("solid", start_color="b7e4c7")
            dala_of_column['E'].border = border






        #SaveFile
        output_excel_file.save(self.output_database_path)
        
        



    def upload_frame_proparty(self):
        self.upload_frame_top_margin_Frame = Frame(self.upload_frame,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.upload_frame_top_margin_Frame.pack()

        self.browse_BTN_Frame = Frame(self.upload_frame,bg=self.colorTk['w'],border=0,borderwidth=0,highlightthickness=0)
        self.browse_BTN_Frame.pack()

        self.browse_BTN = Button(self.browse_BTN_Frame,bg=self.colorTk['w'],activebackground=self.colorTk['w'],height=270,border=0,borderwidth=0,highlightthickness=0)
        self.browse_BTN['image'] = imageTK['xls_gray']
        self.browse_BTN.image = imageTK['xls_gray']
        self.browse_BTN['command'] = self.browse_file_by_filedialog
        self.browse_BTN.pack()

        self.browse_LBL = Label(self.upload_frame,text='Browse file',bg=self.colorTk['w'],fg=self.colorTk['w1'],font=('Arial','20','bold'),border=0,borderwidth=0,highlightthickness=0,pady=5)
        self.browse_LBL.pack()

    def index_image(self):
        percentage = (self.total_processed_report / self.total_row) * 100
        self.tk_progress_bar(percentage)
        im =  ImageTk.PhotoImage(image_title(text=f'{str(self.total_processed_report)} / {str(self.total_row)}',height=27, width= 110, fontsize=20,style=3, color='#EEE8A9'))
        self.processed_report_number_display_ILBL['image'] = im
        self.processed_report_number_display_ILBL.image = im

    def current_index(self):
        im = ImageTk.PhotoImage(image_title(text=f'{str(self.selected_report_index)}',height=40, width= 110, fontsize=45,style=3, color='#EEE8A9'))        
        self.original_report_selected_index_ILBL['image'] = im
        self.original_report_selected_index_ILBL.image = im

    def open_xlsx_file(self):
        path  = self.input_database_path
        work_book = openpyxl.load_workbook(path)
        self.xl_sheet_data = work_book.active

        
        if 'id' in str(self.xl_sheet_data.cell(row=1,column=1).value).strip().lower() :
            
            self.xl_sheet_data.delete_rows(1)
            # Save the changes
            if os.path.exists('./.temp.xlsx'): os.remove('./.temp.xlsx')
            work_book.save('./.temp.xlsx')

            work_book = openpyxl.load_workbook('./.temp.xlsx')
            self.xl_sheet_data = work_book.active


        self.total_row = self.xl_sheet_data.max_row
        
        self.root_data_dict = {}
        self.root_data_process_dict = {}
        for each_row in range(1,self.total_row+1):
            
            self.root_data_process_dict[each_row] = False
            self.root_data_dict[each_row] = {
                'report_id' : str(self.xl_sheet_data.cell(row=int(each_row),column=1).value),
                'original_report' : str(self.xl_sheet_data.cell(row=int(each_row),column=2).value),
                'data_curation' : '',
                'standardize_report' : '',
                'key_findings' : '',
                'data_curation_command' : '',
                'standardize_report_command' : '',
                'key_findings_command' : '',
   
            }
        self.total_row = self.total_row

        # Hide upload Frame
        self.upload_frame.pack_forget()
        #show Working Frame
        self.working__frame_proparty()
        self.working_frame.pack(anchor='w')
        # self.WINDOW.after(1)
        

        #
        self.selected_report_index = 1
        self.current_index()
        
        #
        self.total_processed_report = 0
        self.index_image()

        # Insert value
        self.shift_report()


        



    

    def browse_file_by_filedialog(self):
        self.xl_file_location = filedialog.askopenfilename(initialdir='./',title='Browse input Excel file',filetypes=(('Excel file','*.xlsx'),))
        if self.xl_file_location != None:
            self.xl_file_location = self.xl_file_location.strip()
            if self.xl_file_location.lower().endswith('.xlsx') and os.path.isfile(self.xl_file_location):
                # Import animation
                self.browse_BTN['command'] = ''
                self.browse_BTN['relief'] = SUNKEN
                self.import_animation = gif(self.browse_BTN,image='res/xls_loading_green.gif',speed=60)
                self.import_animation.start()
                self.browse_LBL['fg'] = self.colorTk['g']
                self.import_lbl_animation()

                # self.WINDOW.after(2500)

                
                

    def tk_progress_bar(self,percent):
        image = ImageTk.PhotoImage(progress_image(percentage=percent))
        self.progress_bar_ILB['image'] = image
        self.progress_bar_ILB.image = image
          
    def import_lbl_animation(self,current=1,indent=0,limit=10):
        indent += 1
        if indent == 4 : indent = 0
        base_text = 'Importing'+'.'*indent

        self.browse_LBL['text'] = base_text

        if current <= limit:
            self.WINDOW.after(250,self.import_lbl_animation,current+1,indent,limit)
        else:
            self.import_animation.stop()
            # self.WINDOW.after(1000)
            self.browse_LBL['text'] = 'Imported'

            self.browse_BTN.destroy()

            self.browse_BTN = Button(self.browse_BTN_Frame,bg=self.colorTk['w'],activebackground=self.colorTk['w'],height=270,border=0,borderwidth=0,highlightthickness=0)
            self.browse_BTN['image'] = imageTK['xls']
            self.browse_BTN.image = imageTK['xls']
            self.browse_BTN['relief'] = SUNKEN
            self.browse_BTN.pack()


            # get data from
            self.input_database_path = self.xl_file_location
            self.output_database_path = "/".join(self.input_database_path.split('/')[:-1])+'/output_gpt_report.xlsx'
            self.log_database_path = "/".join(self.input_database_path.split('/')[:-1])+'/gpt_report.log'


            Thread(target=self.open_xlsx_file).start()
            self.WINDOW.after(2000)

            # self.open_xlsx_file()

    def on_window_resize_EL(self,event):
        self.on_window_resize()

    def on_window_resize(self):

        # if self.WINDOW.winfo_height() != self.WINDOW_HEIGHT or self.WINDOW.winfo_width() != self.WINDOW_WIDTH:
        self.WINDOW_HEIGHT = self.WINDOW.winfo_height()
        self.WINDOW_WIDTH = self.WINDOW.winfo_width()



        try: self.upload_frame_top_margin_Frame['height'] = (self.WINDOW_HEIGHT-400)/2
        except: pass
        try:
            # self.log_canvas_height = 80
            comment_canvas_height = 30+20*3+32*3+30+65+30
            self.base_text_canvas['height'] = self.WINDOW_HEIGHT-110-30-comment_canvas_height-self.log_canvas_height-50
            self.base_text_canvas['width'] = int(self.WINDOW_WIDTH)/2
            self.commands_canvas['height'] = comment_canvas_height
            self.commands_canvas['width'] = int(self.WINDOW_WIDTH)/2

            self.file_location_display_frame['width'] = int(self.WINDOW_WIDTH)
            # self.holding_canvas_base_text['height'] = self.WINDOW_HEIGHT-100-30
            # self.holding_canvas_base_text['width'] = int(self.WINDOW_WIDTH)/2
            # self.middle_left_canvas['height'] = self.WINDOW_HEIGHT-100-30
            # self.middle_left_canvas['width'] = int(self.WINDOW_WIDTH)/2

            self.footer_frame_canvas['width'] = int(self.WINDOW_WIDTH)
            self.footer_button_canvas['width'] = int(self.WINDOW_WIDTH)

            # middle_right_canvas

            

            self.standarize_report_text_canvas['height'] = ((self.WINDOW_HEIGHT-110)/3)*2-30-self.log_canvas_height/2-50/2
            self.standarize_report_text_canvas['width'] = int(self.WINDOW_WIDTH)/2
            self.Standarize_report_LOODING_canvas['height'] = ((self.WINDOW_HEIGHT-110)/3)*2-30-self.log_canvas_height/2-50/2
            self.Standarize_report_LOODING_canvas['width'] = int(self.WINDOW_WIDTH)/2-30
            self.standarize_report_loading_ILBL['height'] = ((self.WINDOW_HEIGHT-110)/3)*2-30-self.log_canvas_height/2-50/2
            self.standarize_report_loading_ILBL['width'] = int(self.WINDOW_WIDTH)/2-30
            # self.standarize_report_loading_ILBL
            # self.Standarize_report_LBL_canvas['height'] = (self.WINDOW_HEIGHT-100)/2-30
            # self.Standarize_report_LBL_canvas['width'] = int(self.WINDOW_WIDTH)/2
            # self.standarize_report_loading_ILBL['height'] = (self.WINDOW_HEIGHT-100)/2-30
            # self.standarize_report_loading_ILBL['width'] = int(self.WINDOW_WIDTH)/2
            
            

            self.key_findingds_text_canvas['height'] = ((self.WINDOW_HEIGHT-110)/3)-30-self.log_canvas_height/2-50/2
            self.key_findingds_text_canvas['width'] = int(self.WINDOW_WIDTH)/2
            self.key_findingds_LOODING_canvas['height'] = ((self.WINDOW_HEIGHT-110)/3)-30-self.log_canvas_height/2-50/2
            self.key_findingds_LOODING_canvas['width'] = int(self.WINDOW_WIDTH)/2
            self.key_findings_loading_ILBL['height'] = ((self.WINDOW_HEIGHT-110)/3)-30-self.log_canvas_height/2-50/2
            self.key_findings_loading_ILBL['width'] = int(self.WINDOW_WIDTH)/2

            # self.key_findings_summary_report_LBL_canvas['height'] = (((self.WINDOW_HEIGHT-100)/2)/4)-30
            # self.key_findings_summary_report_LBL_canvas['width'] = int(self.WINDOW_WIDTH)/2
            # self.summary_report_loading_ILBL['height'] = (((self.WINDOW_HEIGHT-100)/2)/4)-30
            # self.summary_report_loading_ILBL['width'] = int(self.WINDOW_WIDTH)/2
            
            
            

            # self.key_findings_summary_canvas['height'] = (((self.WINDOW_HEIGHT-100)/2)/4)*3-30
            # self.key_findings_summary_canvas['width'] = int(self.WINDOW_WIDTH)/2
            # # self.summary_report_LBL_canvas['height'] = (((self.WINDOW_HEIGHT-100)/2)/4)*3-30
            # # self.summary_report_LBL_canvas['width'] = int(self.WINDOW_WIDTH)/2
            # # self.key_findings_loading_ILBL['height'] = (((self.WINDOW_HEIGHT-100)/2)/4)*3-30
            # self.key_findings_loading_ILBL['width'] = int(self.WINDOW_WIDTH)/2


            self.log_canvas['width'] = int(self.WINDOW_WIDTH)/2-60
            self.log_canvas['height'] = self.log_canvas_height

            
            

            



            

            pass
        except: pass
        # print(self.WINDOW.winfo_height())




    

def center_window(window,width=0,height=0):
    ws = window.winfo_screenwidth()
    hs = window.winfo_screenheight()
    x = (ws/2) - (width/2)
    y = (hs/2) - (height/2)
    window.geometry('%dx%d+%d+%d' % (width, height, x, y))





imageTK = {
    'xls' : ImageTk.PhotoImage(Image.open('res/xls.png').resize((256,256),Image.LANCZOS)),
    'xls_gray' : ImageTk.PhotoImage(Image.open('res/xls_gray.png').resize((256,256),Image.LANCZOS)),
    'generate' : ImageTk.PhotoImage(Image.open('res/generate.png').resize((45,45),Image.LANCZOS)),
    'left' : ImageTk.PhotoImage(Image.open('res/left.png').resize((36,36),Image.LANCZOS)),
    'left_gray' : ImageTk.PhotoImage(Image.open('res/left_gray.png').resize((36,36),Image.LANCZOS)),
    'right' : ImageTk.PhotoImage(Image.open('res/right.png').resize((36,36),Image.LANCZOS)),
    'right_gray' : ImageTk.PhotoImage(Image.open('res/right_gray.png').resize((36,36),Image.LANCZOS)),
    'terminal' : ImageTk.PhotoImage(Image.open('res/terminal.png').resize((24,24),Image.LANCZOS)),
    'settings' : ImageTk.PhotoImage(Image.open('res/settings.png').resize((20,20),Image.LANCZOS)),
    'index_select' : ImageTk.PhotoImage(Image.open('res/down_arrow.png').resize((16,8),Image.LANCZOS)),
    'dot_margin' : ImageTk.PhotoImage(Image.open('res/down_arrow.png').resize((1,1),Image.LANCZOS)),
    'api_key' : ImageTk.PhotoImage(Image.open('res/api_key.png').resize((24,24),Image.LANCZOS)),
    'ok' : ImageTk.PhotoImage(Image.open('res/ok.png').resize((40,40),Image.LANCZOS)),
    'xls_24' : ImageTk.PhotoImage(Image.open('res/xls.png').resize((24,24),Image.LANCZOS)),
    'log_24' : ImageTk.PhotoImage(Image.open('res/log.png').resize((24,24),Image.LANCZOS)),

    'std_report' : ImageTk.PhotoImage(image_title(text='Standardize Report',width=200-40)),
    'key_findings' : ImageTk.PhotoImage(image_title(text='Key Findings',width=150-40)),
    'summary_report' : ImageTk.PhotoImage(image_title(text='Summary',width=107-40)),
    'original_report' : ImageTk.PhotoImage(image_title(text='Original report',width=150-13)),
    'list_of_commands' : ImageTk.PhotoImage(image_title(text='List of commands',width=145)),
    'command_1' : ImageTk.PhotoImage(image_title(text='Data curation command',width=170-1,height=30,fontsize=15,color='#324B4B')),
    'command_2' : ImageTk.PhotoImage(image_title(text='Key finding command',width=155-1,height=30,fontsize=15,color='#324B4B')),
    'command_3' : ImageTk.PhotoImage(image_title(text='Report standardization command',width=242,height=30,fontsize=15,color='#324B4B')),
    'run_report' : ImageTk.PhotoImage(image_title(text='Run all reports',width=140,height=20,fontsize=15,color='#324B4B')),
    'run_report2' : ImageTk.PhotoImage(image_title(text='Run all reports',width=140,height=20,fontsize=15,color='#FF7F00')),
    'input_file' : ImageTk.PhotoImage(image_title(text='Input file',width=81,height=20,fontsize=15,color='#324B4B')),
    'output_file' : ImageTk.PhotoImage(image_title(text='Output file',width=90,height=20,fontsize=15,color='#324B4B')),
    'log_file' : ImageTk.PhotoImage(image_title(text='Log file',width=68,height=20,fontsize=15,color='#324B4B')),
}

batch_run_boolVar = BooleanVar()
batch_run_boolVar.set(0)

WINDOW.title('Report Standardizer')
# WINDOW.geometry('1200x750')
center_window(WINDOW,width=1200,height=750)
Evaluation(WINDOW,imageTK)

app_icon = Image.open("res/logo.png")
app_icon = ImageTk.PhotoImage(app_icon.resize((16,16),Image.LANCZOS))
WINDOW.iconphoto(False, app_icon)
WINDOW.mainloop()